# -*- coding: utf8 -*-
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from goods import Goods

#先输入p0--全国平均单价
def inputP0(P0,HSCODE_quanguo):   
    wb=load_workbook(filename = 'data//quanguo.xlsx', use_iterators=True)
    sheet_names= wb.get_sheet_names()
    #数据表必须是数据文件第一个表
    ws=wb.get_sheet_by_name(sheet_names[0])     
    d={}
    for row in ws.iter_rows():
        #读入全国平均单价
        tmp=row[P0].internal_value
        if str(tmp).isdigit():
            tmp=float(tmp)
        #创建聚合的价格数据              
        result=Goods(0.0,0.0,tmp)
        #与8位税号建立关联关系                    
        d[(str(row[HSCODE_quanguo].internal_value)).strip()[0:8]]=result    
    return d

#输入通关数据
def inputP1(d,HSCODE_tongguan,Qty,Amount):             
    dirs=os.listdir('data//')
    print dirs[0]
    for filename in dirs:
        names=filename.split('.')
        if names[0]!='quanguo' and names[1]=='xlsx':
            wb=load_workbook(filename= 'data//'+filename,use_iterators=True)
            sheet_names=wb.get_sheet_names()
            ws=wb.get_sheet_by_name(sheet_names[0])          
            for row in ws.iter_rows(row_offset=1,):
                #不同数据文件需要调整行号
                hscode=unicode(row[HSCODE_tongguan].internal_value)[0:8]
                #如果全国平均单价没有此类，则舍弃该数据             
                if hscode in d:              
                    d[hscode].ttl+=float(row[Qty].internal_value)
                    d[hscode].amout+=float(row[Amount].internal_value)
                    # 引用通关数据的统计美元值进行计算
                    # d[hscode].huilv=float(row[19].internal_value)                
    return d

#奇异点判断函数
def flag(v):                     
    if v.p1==0:
        result=False
    else:
        result=True
    return result

#计算指标函数
def calc(d):                      
    fenzi=0.0
    fenmu=0.0
    wb=Workbook(optimized_write= True)
    ws=wb.create_sheet()
    ws.append(['HS_code','影响度'])    
    for k,v in d.iteritems():
        v.setP1()
        #引入排除奇异点函数
        if flag(v):                         
            fenzi+=v.p1*v.ttl
            fenmu+=v.p0*v.ttl
    print "fenzi:%F; fenmu:%F"%(fenzi,fenmu)
    #计算每个税号的影响度
    for k,v in d.iteritems():
        if flag(v):            
            v.affect=v.ttl*(v.p1-v.p0)/fenmu*100.0
            ws.append([k,v.affect])   
    wb.save('affect.xlsx')
    return fenzi/fenmu*100.0
            

def main():
    config=[]
    with open("config.ini") as f:
        for line in f:
            config.append(int(line.split(':')[1])-1)
        f.close()
    d=inputP0(config[0],config[1])
    d=inputP1(d,config[2],config[3],config[4])       
    print calc(d)
    
if __name__=='__main__':
    main()
